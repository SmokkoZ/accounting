"""
Ledger Export Service

Handles exporting the full ledger to styled Excel workbooks for audit
and operator review while maintaining strict validation guarantees.
"""

from __future__ import annotations

import csv
from datetime import datetime
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from xlsxwriter.utility import xl_col_to_name
import structlog
from src.core.database import get_db_connection

logger = structlog.get_logger()


@dataclass
class LedgerExportResult:
    """Metadata returned after a ledger export completes."""

    file_path: str
    row_count: int
    file_size_bytes: int
    associate_id: Optional[int]
    scope_label: str


class LedgerExportService:
    """Service for exporting ledger entries to Excel workbook format."""

    def __init__(self, export_dir: str = "data/exports"):
        base_dir = Path(export_dir)
        self.export_dir = base_dir / "ledger"
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export_full_ledger(self, associate_id: Optional[int] = None) -> LedgerExportResult:
        """
        Export complete ledger to a styled Excel workbook with optional associate filtering.
        
        Returns:
            LedgerExportResult describing the generated workbook
            
        Raises:
            Exception: If export fails for any reason
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        associate_alias: Optional[str] = None
        associate_currency: Optional[str] = None
        scope_label: Optional[str] = None
        file_path: Optional[Path] = None

        conn = None
        try:
            conn = get_db_connection()
            if associate_id is not None:
                associate_alias, associate_currency = self._get_associate_details(
                    conn, associate_id
                )
                if associate_alias is None:
                    raise ValueError(f"Associate {associate_id} not found for export")
                scope_label = associate_alias or f"Associate #{associate_id}"
            else:
                associate_alias = "all"
                associate_currency = "MULTI"
                scope_label = "All Associates"

            filename = self._build_filename(
                timestamp, associate_alias, associate_currency
            )
            file_path = self._make_unique_path(filename)

            logger.info(
                "starting_ledger_export",
                filename=filename,
                file_path=str(file_path),
                associate_id=associate_id,
            )

            row_count = self._export_ledger_to_excel(conn, file_path, associate_id)

            # Validate export
            self._validate_export(file_path, row_count)

            logger.info(
                "ledger_export_completed",
                file_path=str(file_path),
                row_count=row_count,
                associate_id=associate_id,
            )
            size_bytes = file_path.stat().st_size if file_path and file_path.exists() else 0
            return LedgerExportResult(
                file_path=str(file_path),
                row_count=row_count,
                file_size_bytes=size_bytes,
                associate_id=associate_id,
                scope_label=scope_label or "All Associates",
            )

        except Exception as e:
            logger.error(
                "ledger_export_failed",
                error=str(e),
                associate_id=associate_id,
                file_path=str(file_path) if file_path else None,
            )
            raise
        finally:
            if conn is not None:
                conn.close()

    def _export_ledger_to_excel(
        self, conn, file_path: Path, associate_id: Optional[int] = None
    ) -> int:
        """Execute ledger query and write to Excel file."""
        bet_columns = self._get_table_columns(conn, "bets")
        selection_text_expr = self._column_expr("b", "selection_text", bet_columns, "selection_text")
        selection_expr = self._column_expr("b", "selection", bet_columns, "selection")
        market_code_expr = self._column_expr("b", "market_code", bet_columns, "market_code")
        side_expr = self._column_expr("b", "side", bet_columns, "side")
        line_value_expr = self._column_expr("b", "line_value", bet_columns, "line_value")

        base_query = f"""
            SELECT
                le.id AS entry_id,
                le.type AS entry_type,
                a.display_alias AS associate_alias,
                bk.bookmaker_name,
                {selection_text_expr},
                {selection_expr},
                {market_code_expr},
                {side_expr},
                {line_value_expr},
                cm.description AS market_description,
                ce.normalized_event_name AS canonical_event_name,
                le.amount_native,
                le.native_currency,
                le.fx_rate_snapshot,
                le.amount_eur,
                le.settlement_state,
                le.principal_returned_eur,
                le.per_surebet_share_eur,
                le.surebet_id,
                le.bet_id,
                le.settlement_batch_id,
                le.created_at_utc,
                le.created_by,
                le.note
            FROM ledger_entries le
            JOIN associates a ON le.associate_id = a.id
            LEFT JOIN bookmakers bk ON le.bookmaker_id = bk.id
            LEFT JOIN bets b ON le.bet_id = b.id
            LEFT JOIN canonical_events ce ON b.canonical_event_id = ce.id
            LEFT JOIN canonical_markets cm ON b.canonical_market_id = cm.id
        """

        params: Tuple = ()
        if associate_id is not None:
            base_query += " WHERE le.associate_id = ?"
            params = (associate_id,)

        query = f"{base_query} ORDER BY le.created_at_utc ASC"

        cursor = conn.cursor()
        cursor.execute(query, params)

        # Define workbook columns
        fieldnames = [
            "created_at_utc",
            "entry_type",
            "associate_alias",
            "bookmaker_name",
            "event_name",
            "market_selection",
            "settlement_state",
            "native_currency",
            "amount_native",
            "principal_returned_native",
            "note",
            "amount_eur",
            "principal_returned_eur",
            "per_surebet_share_eur",
            "entry_id",
            "surebet_id",
            "bet_id",
            "settlement_batch_id",
            "fx_rate_snapshot",
            "created_by",
        ]

        rows = cursor.fetchall()
        formatted_rows = [self._format_row_for_export(dict(row)) for row in rows]
        row_count = len(formatted_rows)

        if formatted_rows:
            dataframe = pd.DataFrame(formatted_rows, columns=fieldnames)
        else:
            dataframe = pd.DataFrame(columns=fieldnames)

        dataframe = dataframe[fieldnames]
        export_df = dataframe.where(pd.notnull(dataframe), None)
        column_widths = self._compute_column_widths(dataframe, fieldnames)

        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            export_df.to_excel(writer, sheet_name="Ledger", index=False)
            workbook = writer.book
            worksheet = writer.sheets["Ledger"]

            header_format = workbook.add_format(
                {
                    "bold": True,
                    "bg_color": "#E5ECFF",
                    "border": 1,
                    "text_wrap": False,
                }
            )
            for col_num, header in enumerate(fieldnames):
                worksheet.write(0, col_num, header, header_format)

            worksheet.freeze_panes(1, 0)
            worksheet.autofilter(0, 0, max(row_count, 1), len(fieldnames) - 1)

            format_map = self._build_column_formats(workbook, fieldnames)
            for idx, column in enumerate(fieldnames):
                worksheet.set_column(
                    idx,
                    idx,
                    column_widths.get(column, len(column) + 2),
                    format_map.get(column),
                )

            if row_count > 0 and "entry_type" in dataframe.columns:
                entry_col_letter = xl_col_to_name(fieldnames.index("entry_type"))
                deposit_format = workbook.add_format({"bg_color": "#E6F4EA"})
                withdrawal_format = workbook.add_format({"bg_color": "#FDECEA"})
                worksheet.conditional_format(
                    1,
                    0,
                    row_count,
                    len(fieldnames) - 1,
                    {
                        "type": "formula",
                        "criteria": f'=${entry_col_letter}2="DEPOSIT"',
                        "format": deposit_format,
                    },
                )
                worksheet.conditional_format(
                    1,
                    0,
                    row_count,
                    len(fieldnames) - 1,
                    {
                        "type": "formula",
                        "criteria": f'=${entry_col_letter}2="WITHDRAWAL"',
                        "format": withdrawal_format,
                    },
                )

        return row_count

    def _compute_column_widths(self, dataframe: pd.DataFrame, columns: List[str]) -> Dict[str, float]:
        """Calculate a reasonable width for each column based on cell content."""
        widths: Dict[str, float] = {}
        for column in columns:
            if column in dataframe:
                series = dataframe[column]
            else:
                series = pd.Series(dtype=object)

            if series.empty:
                max_length = len(column)
            else:
                display_series = series.fillna("").astype(str)
                max_length = int(display_series.map(len).max())
            widths[column] = float(min(max(max_length, len(column)) + 2, 60))
        return widths

    def _build_column_formats(self, workbook, columns: List[str]) -> Dict[str, object]:
        """Map ledger columns to Excel numeric formats."""
        currency_columns = {
            "amount_native",
            "principal_returned_native",
            "amount_eur",
            "principal_returned_eur",
            "per_surebet_share_eur",
        }
        fx_columns = {"fx_rate_snapshot"}
        integer_columns = {"entry_id", "surebet_id", "bet_id"}

        currency_format = workbook.add_format({"num_format": "#,##0.00"})
        fx_format = workbook.add_format({"num_format": "0.0000"})
        integer_format = workbook.add_format({"num_format": "0"})

        format_map: Dict[str, object] = {}
        for column in columns:
            if column in currency_columns:
                format_map[column] = currency_format
            elif column in fx_columns:
                format_map[column] = fx_format
            elif column in integer_columns:
                format_map[column] = integer_format
            else:
                format_map[column] = None
        return format_map

    def _format_row_for_export(self, row: Dict) -> Dict[str, object]:
        """Format a database row for Excel output."""
        formatted: Dict[str, object] = {}

        event_name = (
            row.get("canonical_event_name")
            or row.get("selection_text")
            or row.get("selection")
            or ""
        )
        market_label = row.get("market_description") or row.get("market_code") or ""
        side = row.get("side")
        line_value = row.get("line_value")
        if side:
            market_label = f"{market_label} - {side}" if market_label else side
        if line_value:
            market_label = f"{market_label} ({line_value})" if market_label else str(line_value)

        principal_native: Optional[float] = None
        principal_eur = row.get("principal_returned_eur")
        fx_rate = row.get("fx_rate_snapshot")
        if principal_eur and fx_rate:
            try:
                fx_decimal = Decimal(fx_rate)
                if fx_decimal != 0:
                    native_value = Decimal(principal_eur) / fx_decimal
                    principal_native = float(native_value.quantize(Decimal("0.01")))
            except (InvalidOperation, ZeroDivisionError):
                principal_native = None

        row["event_name"] = event_name
        row["market_selection"] = market_label
        row["principal_returned_native"] = principal_native
        row["created_at_utc"] = self._format_created_date(row.get("created_at_utc"))
        row["note"] = row.get("note") or ""
        row["bookmaker_name"] = row.get("bookmaker_name") or ""
        row["associate_alias"] = row.get("associate_alias") or ""
        row["native_currency"] = row.get("native_currency") or ""
        row["settlement_state"] = row.get("settlement_state") or ""

        decimal_fields = {
            "amount_native",
            "fx_rate_snapshot",
            "amount_eur",
            "principal_returned_eur",
            "principal_returned_native",
            "per_surebet_share_eur",
        }
        integer_fields = {"entry_id", "surebet_id", "bet_id"}
        helper_fields = {
            "canonical_event_name",
            "selection_text",
            "selection",
            "market_code",
            "market_description",
            "side",
            "line_value",
        }

        for key, value in row.items():
            if key in helper_fields:
                continue
            if key in decimal_fields:
                formatted[key] = self._to_float(value)
            elif key in integer_fields:
                formatted[key] = self._to_int(value)
            else:
                formatted[key] = "" if value in (None, "") else str(value)

        return formatted

    def _to_float(self, value: object) -> Optional[float]:
        """Safely convert value to float."""
        if value in (None, ""):
            return None
        try:
            return float(Decimal(str(value)))
        except (InvalidOperation, TypeError, ValueError):
            return None

    def _to_int(self, value: object) -> Optional[int]:
        """Safely convert value to integer."""
        if value in (None, ""):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            try:
                return int(float(value))
            except (TypeError, ValueError):
                return None

    def _format_created_date(self, value: Optional[str]) -> str:
        """Format created_at date to DD/MM/YYYY."""
        if not value:
            return ""
        try:
            date_part = value[:10]
            dt = datetime.strptime(date_part, "%Y-%m-%d")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            try:
                cleaned = value.replace("Z", "+00:00")
                dt = datetime.fromisoformat(cleaned)
                return dt.strftime("%d/%m/%Y")
            except ValueError:
                return value

    def _validate_export(self, file_path: Path, expected_row_count: int) -> None:
        """Validate that export was successful."""
        if not file_path.exists():
            raise FileNotFoundError(f"Export file not created: {file_path}")

        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            workbook = load_workbook(file_path, data_only=True)
            try:
                worksheet = workbook.active
                actual_row_count = sum(
                    1
                    for row in worksheet.iter_rows(min_row=2, values_only=True)
                    if any(cell not in (None, "") for cell in row)
                )
            finally:
                workbook.close()
        else:
            with open(file_path, "r", encoding="utf-8") as csvfile:
                reader = csv.DictReader(csvfile)
                actual_row_count = sum(1 for _ in reader)

        if actual_row_count != expected_row_count:
            raise ValueError(
                f"Row count mismatch: expected {expected_row_count}, "
                f"found {actual_row_count} in export"
            )

        logger.info("export_validation_passed", file_path=str(file_path), row_count=actual_row_count)

    def get_export_history(self, limit: int = 10) -> List[Dict]:
        """
        Get list of recent export files with metadata.
        
        Args:
            limit: Maximum number of files to return
            
        Returns:
            List of dictionaries with file metadata
        """
        exports = []
        
        if not self.export_dir.exists():
            return exports

        # Collect ledger export files under both new and legacy naming schemes
        export_candidates: Set[Path] = set()
        export_candidates.update(self.export_dir.glob("*_ledger.xlsx"))
        export_candidates.update(self.export_dir.glob("ledger_*.xlsx"))
        export_candidates.update(self.export_dir.glob("*_ledger.csv"))
        export_candidates.update(self.export_dir.glob("ledger_*.csv"))

        export_files = sorted(
            export_candidates,
            key=lambda f: f.stat().st_mtime,
            reverse=True
        )

        alias_cache: Dict[int, Optional[str]] = {}
        conn = None
        attempted_conn = False

        for file_path in export_files[:limit]:
            try:
                stat = file_path.stat()
                scope_meta = self._parse_export_filename(file_path.name)
                
                # Count rows in file
                if file_path.suffix.lower() == ".xlsx":
                    try:
                        dataframe = pd.read_excel(file_path, engine="openpyxl")
                        row_count = len(dataframe.index)
                    except Exception:
                        row_count = 0
                else:
                    with open(file_path, "r", encoding="utf-8") as csvfile:
                        reader = csv.DictReader(csvfile)
                        row_count = sum(1 for _ in reader)

                associate_alias = None
                alias_slug = scope_meta.get("alias_slug")
                associate_id = scope_meta.get("associate_id")
                if associate_id is not None:
                    if conn is None and not attempted_conn:
                        try:
                            conn = get_db_connection()
                        except Exception:
                            conn = None
                        finally:
                            attempted_conn = True

                    associate_alias = self._get_cached_alias(conn, associate_id, alias_cache)
                elif scope_meta.get("scope") == "associate" and alias_slug:
                    associate_alias = alias_slug.replace("-", " ").title()

                exports.append({
                    "filename": file_path.name,
                    "file_path": str(file_path),
                    "file_size": stat.st_size,
                    "row_count": row_count,
                    "created_time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                    "scope": scope_meta.get("scope"),
                    "associate_id": associate_id,
                    "associate_alias": associate_alias,
                    "alias_slug": alias_slug,
                    "currency": scope_meta.get("currency"),
                })
            except Exception as e:
                logger.warning("failed_to_read_export_metadata", file=str(file_path), error=str(e))
                continue

        if conn is not None:
            conn.close()

        return exports

    def get_file_size_display(self, size_bytes: int) -> str:
        """Convert file size to human-readable format."""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.1f} TB"

    def _build_filename(
        self,
        timestamp: str,
        associate_alias: Optional[str],
        associate_currency: Optional[str],
    ) -> str:
        alias_slug = self._slugify(associate_alias) if associate_alias else "all"
        currency_slug = (associate_currency or "MULTI").upper()
        try:
            dt = datetime.strptime(timestamp, "%Y%m%d_%H%M%S_%f")
        except ValueError:
            date_part = timestamp.split("_", 1)[0]
            dt = datetime.strptime(date_part, "%Y%m%d")
        date_str = dt.strftime("%d-%m-%Y")
        return f"{alias_slug}_{currency_slug}_{date_str}_ledger.xlsx"

    def _make_unique_path(self, base_filename: str) -> Path:
        """Ensure the export filename is unique within the export directory."""
        path = self.export_dir / base_filename
        if not path.exists():
            return path

        stem = path.stem
        suffix = path.suffix
        counter = 1
        while True:
            candidate = self.export_dir / f"{stem}_{counter}{suffix}"
            if not candidate.exists():
                return candidate
            counter += 1

    def _slugify(self, value: Optional[str]) -> str:
        if not value:
            return "associate"
        slug = re.sub(r"[^A-Za-z0-9]+", "-", value.strip()).strip("-").lower()
        return slug or "associate"

    def _get_associate_details(
        self, conn, associate_id: int
    ) -> Tuple[Optional[str], Optional[str]]:
        cursor = conn.execute(
            "SELECT display_alias, home_currency FROM associates WHERE id = ?",
            (associate_id,),
        )
        row = cursor.fetchone()
        if not row:
            return None, None
        return row["display_alias"], row["home_currency"]

    def _get_associate_alias(self, conn, associate_id: int) -> Optional[str]:
        alias, _ = self._get_associate_details(conn, associate_id)
        return alias

    def _parse_export_filename(self, filename: str) -> Dict[str, Optional[object]]:
        timestamp_pattern = r"[0-9]{8}_[0-9]{6}(?:_[0-9]{6})?"
        date_pattern = r"[0-9]{2}-[0-9]{2}-[0-9]{4}"
        extension_pattern = r"(?:csv|xlsx)"
        new_pattern = re.compile(
            rf"^([A-Za-z0-9\-]+)_([A-Z]+)_({date_pattern})(?:_(\d+))?_ledger\.{extension_pattern}$"
        )
        assoc_pattern = re.compile(
            rf"^ledger_assoc-(\d+)-([A-Za-z0-9\-]+)_({timestamp_pattern})\.{extension_pattern}$"
        )
        full_pattern = re.compile(rf"^ledger_({timestamp_pattern})\.{extension_pattern}$")

        new_match = new_pattern.match(filename)
        if new_match:
            alias_slug = new_match.group(1)
            scope = "all" if alias_slug == "all" else "associate"
            return {
                "scope": scope,
                "associate_id": None,
                "alias_slug": alias_slug,
                "currency": new_match.group(2),
                "suffix": new_match.group(4),
            }

        assoc_match = assoc_pattern.match(filename)
        if assoc_match:
            return {
                "scope": "associate",
                "associate_id": int(assoc_match.group(1)),
                "alias_slug": assoc_match.group(2),
                "currency": None,
                "suffix": None,
            }

        if full_pattern.match(filename):
            return {
                "scope": "all",
                "associate_id": None,
                "alias_slug": None,
                "currency": None,
                "suffix": None,
            }

        return {"scope": "unknown", "associate_id": None, "alias_slug": None, "currency": None, "suffix": None}

    def _get_cached_alias(
        self,
        conn,
        associate_id: int,
        cache: Dict[int, Optional[str]],
    ) -> Optional[str]:
        if associate_id in cache:
            return cache[associate_id]

        if conn is None:
            cache[associate_id] = None
            return None

        alias = self._get_associate_alias(conn, associate_id)
        cache[associate_id] = alias
        return alias

    def _get_table_columns(self, conn, table_name: str) -> Set[str]:
        """Return set of column names for a table."""
        cursor = conn.execute(f"PRAGMA table_info({table_name})")
        return {row["name"] for row in cursor.fetchall()}

    def _column_expr(self, table_alias: str, column: str, columns: Set[str], alias: str) -> str:
        """Return a safe column reference or NULL if column missing."""
        if column in columns:
            return f"{table_alias}.{column} AS {alias}"
        return f"NULL AS {alias}"
