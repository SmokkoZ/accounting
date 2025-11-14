"""
Integration tests for Statement generation workflow.

Tests complete flow from UI to database with realistic data scenarios.
"""

import io
import importlib
import os
import sqlite3
import tempfile
from datetime import datetime
from decimal import Decimal
from unittest.mock import patch, Mock

import pytest
from openpyxl import load_workbook

from src.core.schema import create_schema
from src.services.statement_service import (
    BookmakerStatementRow,
    StatementCalculations,
    StatementService,
)


@pytest.fixture
def service():
    """Create StatementService instance."""
    return StatementService()


@pytest.fixture
def temp_statement_db():
    """Temporary sqlite database path for settlement workflow tests."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
    tmp.close()
    try:
        yield tmp.name
    finally:
        os.unlink(tmp.name)


@pytest.fixture
def sample_associate_data():
    """Sample associate data for testing."""
    return {
        "id": 1,
        "display_alias": "Test Associate",
        "home_currency": "EUR",
    }


@pytest.fixture
def sample_ledger_entries():
    """Sample ledger entries for testing."""
    return [
        # Deposits
        {
            "id": 1,
            "type": "DEPOSIT",
            "associate_id": 1,
            "amount_eur": "1000.00",
            "native_currency": "EUR",
            "amount_native": "1000.00",
            "fx_rate_snapshot": "1.0",
            "settlement_state": None,
            "principal_returned_eur": None,
            "per_surebet_share_eur": None,
            "surebet_id": None,
            "bet_id": None,
            "created_at_utc": "2025-10-15T10:00:00Z",
            "note": "Initial deposit"
        },
        # Bet result - WON
        {
            "id": 2,
            "type": "BET_RESULT",
            "associate_id": 1,
            "amount_eur": "-100.00",
            "native_currency": "EUR",
            "amount_native": "-100.00",
            "fx_rate_snapshot": "1.0",
            "settlement_state": "WON",
            "principal_returned_eur": "100.00",
            "per_surebet_share_eur": "50.00",
            "surebet_id": 1,
            "bet_id": 1,
            "created_at_utc": "2025-10-20T15:00:00Z",
            "note": "Bet settlement - Won"
        },
        # Withdrawal
        {
            "id": 3,
            "type": "WITHDRAWAL",
            "associate_id": 1,
            "amount_eur": "-200.00",
            "native_currency": "EUR",
            "amount_native": "-200.00",
            "fx_rate_snapshot": "1.0",
            "settlement_state": None,
            "principal_returned_eur": None,
            "per_surebet_share_eur": None,
            "surebet_id": None,
            "bet_id": None,
            "created_at_utc": "2025-10-25T12:00:00Z",
            "note": "Withdrawal"
        },
        # Another bet result - WON
        {
            "id": 4,
            "type": "BET_RESULT",
            "associate_id": 1,
            "amount_eur": "-150.00",
            "native_currency": "EUR",
            "amount_native": "-150.00",
            "fx_rate_snapshot": "1.0",
            "settlement_state": "WON",
            "principal_returned_eur": "150.00",
            "per_surebet_share_eur": "75.00",
            "surebet_id": 2,
            "bet_id": 2,
            "created_at_utc": "2025-10-28T18:00:00Z",
            "note": "Bet settlement - Won"
        }
    ]


@pytest.fixture
def mock_database_with_data(sample_associate_data, sample_ledger_entries):
    """Mock database with sample data."""
    conn = Mock()
    cursor = Mock()
    conn.cursor.return_value = cursor
    conn.close = Mock()
    
    # Mock associate lookup
    def mock_execute(query, params=None):
        normalized = " ".join(query.split())
        if "SELECT display_alias" in normalized and "FROM associates" in normalized:
            cursor.fetchone.return_value = sample_associate_data
            cursor.fetchall.return_value = []
        elif "SUM( CASE WHEN type = 'DEPOSIT'" in normalized and "total_deposits" in normalized:
            cursor.fetchone.return_value = {
                "total_deposits": 1000.0,
                "signed_withdrawals": -200.0,
            }
        elif "principal_returned_eur AS REAL" in normalized:
            cursor.fetchone.return_value = {"should_hold_eur": 275.0}
        elif "profit_before_payout_eur" in normalized:
            cursor.fetchone.return_value = {"profit_before_payout_eur": 125.0}
        elif "SUM(CAST(amount_eur AS REAL)) AS current_holding_eur" in normalized:
            cursor.fetchone.return_value = {"current_holding_eur": 550.0}
        elif "FROM bookmakers" in normalized and "balance_eur" in normalized:
            cursor.fetchone.return_value = None
            cursor.fetchall.return_value = [
                {
                    "bookmaker_name": "Bookie One",
                    "balance_eur": 550.0,
                    "deposits_eur": 1000.0,
                    "withdrawals_eur": 200.0,
                    "balance_native": 550.0,
                    "native_currency": "EUR",
                }
            ]
        elif "SELECT id, type, amount_eur" in normalized:
            cursor.fetchone.return_value = None
            cursor.fetchall.return_value = sample_ledger_entries
        else:
            cursor.fetchone.return_value = None
            cursor.fetchall.return_value = []
    
    cursor.execute.side_effect = mock_execute
    
    return conn, cursor


class TestCompleteStatementGeneration:
    """Test complete statement generation with realistic data."""
    
    def test_profitable_associate_statement(self, service, mock_database_with_data):
        """Test statement generation for profitable associate."""
        conn, cursor = mock_database_with_data
        
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            result = service.generate_statement(1, "2025-10-31T23:59:59Z")
        
        # Verify calculations
        assert result.associate_name == "Test Associate"
        assert result.net_deposits_eur == Decimal("800.00")    # 1000 deposit - 200 withdrawal
        assert result.should_hold_eur == Decimal("275.00")      # 100+50 + 150+75
        assert result.current_holding_eur == Decimal("550.00")   # 1000 - 100 - 200 - 150
        assert result.profit_before_payout_eur == Decimal("125.00")  # 50 + 75
        assert result.raw_profit_eur == Decimal("-525.00")       # 275 - 800 (loss)
        assert result.delta_eur == Decimal("275.00")           # 550 - 275
        
        # Verify database was called correctly
        assert cursor.execute.call_count >= 4  # associate + 3 calculations
    
    def test_loss_associate_statement(self, service):
        """Test statement generation for associate with losses."""
        conn = Mock()
        cursor = Mock()
        conn.cursor.return_value = cursor
        
        # Mock data for loss scenario
        cursor.fetchone.side_effect = [
            {"display_alias": "Loss Associate", "home_currency": "EUR"},
            {"total_deposits": 2000.0, "signed_withdrawals": 0.0},
            {"current_holding_eur": 1100.0},
            {"should_hold_eur": 1200.0},
            {"profit_before_payout_eur": 0.0},
        ]
        cursor.fetchall.return_value = []
        
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            result = service.generate_statement(1, "2025-10-31T23:59:59Z")
        
        assert result.net_deposits_eur == Decimal("2000.00")
        assert result.should_hold_eur == Decimal("1200.00")
        assert result.current_holding_eur == Decimal("1100.00")
        assert result.profit_before_payout_eur == Decimal("0.00")
        assert result.raw_profit_eur == Decimal("-800.00")    # 1200 - 2000 (loss)
        assert result.delta_eur == Decimal("-100.00")        # 1100 - 1200 (short)
    
    def test_balanced_associate_statement(self, service):
        """Test statement generation for balanced associate."""
        conn = Mock()
        cursor = Mock()
        conn.cursor.return_value = cursor
        
        # Mock data for balanced scenario
        cursor.fetchone.side_effect = [
            {"display_alias": "Balanced Associate", "home_currency": "EUR"},
            {"total_deposits": 1000.0, "signed_withdrawals": 0.0},
            {"current_holding_eur": 1000.0},
            {"should_hold_eur": 1000.0},
            {"profit_before_payout_eur": 0.0},
        ]
        cursor.fetchall.return_value = []
        
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            result = service.generate_statement(1, "2025-10-31T23:59:59Z")
        
        assert result.net_deposits_eur == Decimal("1000.00")
        assert result.should_hold_eur == Decimal("1000.00")
        assert result.current_holding_eur == Decimal("1000.00")
        assert result.profit_before_payout_eur == Decimal("0.00")
        assert result.raw_profit_eur == Decimal("0.00")      # 1000 - 1000 (break-even)
        assert result.delta_eur == Decimal("0.00")           # 1000 - 1000 (balanced)


class TestTransactionRetrieval:
    """Test transaction retrieval for export functionality."""
    
    def test_get_associate_transactions_for_export(self, service, mock_database_with_data):
        """Test transaction retrieval for Excel export."""
        conn, cursor = mock_database_with_data
        
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            transactions = service.get_associate_transactions(1, "2025-10-31T23:59:59Z")
        
        assert len(transactions) == 4
        assert transactions[0]["type"] == "DEPOSIT"
        assert transactions[1]["type"] == "BET_RESULT"
        assert transactions[2]["type"] == "WITHDRAWAL"
        assert transactions[3]["type"] == "BET_RESULT"
        
        # Verify cutoff date filtering (all transactions should be before cutoff)
        for transaction in transactions:
            assert transaction["created_at_utc"] <= "2025-10-31T23:59:59Z"
    
    def test_get_associate_transactions_empty_result(self, service):
        """Test transaction retrieval with no matching transactions."""
        conn = Mock()
        cursor = Mock()
        conn.cursor.return_value = cursor
        
        cursor.fetchall.return_value = []
        
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            transactions = service.get_associate_transactions(1, "2025-10-31T23:59:59Z")
        
        assert transactions == []


class TestFormattingIntegration:
    """Test formatting with realistic scenarios."""
    
    def test_partner_facing_section_realistic_profit(self, service):
        """Test partner-facing formatting with realistic profit scenario."""

        calc = StatementCalculations(
            associate_id=1,
            net_deposits_eur=Decimal("5000.00"),
            should_hold_eur=Decimal("7500.00"),
            current_holding_eur=Decimal("7200.00"),
            fair_share_eur=Decimal("1800.00"),
            profit_before_payout_eur=Decimal("1800.00"),
            raw_profit_eur=Decimal("2500.00"),
            delta_eur=Decimal("-300.00"),
            total_deposits_eur=Decimal("5000.00"),
            total_withdrawals_eur=Decimal("0.00"),
            bookmakers=[
                BookmakerStatementRow(
                    bookmaker_name="Profitable Bookie",
                    balance_eur=Decimal("7200.00"),
                    deposits_eur=Decimal("5000.00"),
                    withdrawals_eur=Decimal("0.00"),
                    balance_native=Decimal("7200.00"),
                    native_currency="EUR",
                )
            ],
            associate_name="Profitable Associate",
            home_currency="EUR",
            cutoff_date="2025-10-31T23:59:59Z",
            generated_at="2025-11-05T11:00:00Z"
        )
        
        result = service.format_partner_facing_section(calc)
        
        # Verify realistic profit formatting
        assert result.total_deposits_eur == Decimal("5000.00")
        assert result.total_balance_eur == Decimal("7200.00")
        assert result.bookmakers[0].bookmaker_name == "Profitable Bookie"
    
    def test_internal_section_realistic_scenarios(self, service):
        """Test internal section formatting with realistic scenarios."""
        
        # Test holding more scenario
        calc_holding_more = StatementCalculations(
            associate_id=1,
            net_deposits_eur=Decimal("1000.00"),
            should_hold_eur=Decimal("2000.00"),
            current_holding_eur=Decimal("2500.00"),
            fair_share_eur=Decimal("400.00"),
            profit_before_payout_eur=Decimal("400.00"),
            raw_profit_eur=Decimal("1000.00"),
            delta_eur=Decimal("500.00"),
            total_deposits_eur=Decimal("1000.00"),
            total_withdrawals_eur=Decimal("0.00"),
            bookmakers=[],
            associate_name="Test Associate",
            home_currency="EUR",
            cutoff_date="2025-10-31T23:59:59Z",
            generated_at="2025-11-05T11:00:00Z"
        )
        
        result = service.format_internal_section(calc_holding_more)
        assert "Currently holding: EUR 2,500.00" in result.current_holdings
        assert "Holding more by EUR 500.00" in result.reconciliation_delta
        assert result.delta_indicator == "over"
        
        # Test short scenario
        calc_short = StatementCalculations(
            associate_id=1,
            net_deposits_eur=Decimal("1000.00"),
            should_hold_eur=Decimal("2000.00"),
            current_holding_eur=Decimal("1500.00"),
            fair_share_eur=Decimal("200.00"),
            profit_before_payout_eur=Decimal("200.00"),
            raw_profit_eur=Decimal("1000.00"),
            delta_eur=Decimal("-500.00"),
            total_deposits_eur=Decimal("1000.00"),
            total_withdrawals_eur=Decimal("0.00"),
            bookmakers=[],
            associate_name="Test Associate",
            home_currency="EUR",
            cutoff_date="2025-10-31T23:59:59Z",
            generated_at="2025-11-05T11:00:00Z"
        )
        
        result_short = service.format_internal_section(calc_short)
        assert "Short by EUR 500.00" in result_short.reconciliation_delta
        assert result_short.delta_indicator == "short"


class TestCutoffDateScenarios:
    """Test various cutoff date scenarios."""
    
    def test_month_end_cutoff(self, service):
        """Test statement generation with month-end cutoff."""
        conn = Mock()
        cursor = Mock()
        conn.cursor.return_value = cursor
        
        # Mock data
        cursor.fetchone.side_effect = [
            {"display_alias": "Test Associate", "home_currency": "EUR"},
            {"total_deposits": 1000.0, "signed_withdrawals": 0.0},
            {"current_holding_eur": 1150.0},
            {"should_hold_eur": 1200.0},
            {"profit_before_payout_eur": 0.0},
        ]
        cursor.fetchall.return_value = []
        
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            result = service.generate_statement(1, "2025-10-31T23:59:59Z")
        
        assert result.cutoff_date == "2025-10-31T23:59:59Z"
        
        # Verify cutoff date was used in queries
        execute_calls = cursor.execute.call_args_list
        cutoff_date_found = False
        for call in execute_calls:
            if call[0] and "created_at_utc <= ?" in call[0][0]:
                cutoff_date_found = True
                assert "2025-10-31T23:59:59Z" in call[0][1]
        
        assert cutoff_date_found
    
    def test_mid_month_cutoff(self, service):
        """Test statement generation with mid-month cutoff."""
        conn = Mock()
        cursor = Mock()
        conn.cursor.return_value = cursor
        
        # Mock data
        cursor.fetchone.side_effect = [
            {"display_alias": "Test Associate", "home_currency": "EUR"},
            {"total_deposits": 500.0, "signed_withdrawals": 0.0},
            {"current_holding_eur": 550.0},
            {"should_hold_eur": 600.0},
            {"profit_before_payout_eur": 0.0},
        ]
        cursor.fetchall.return_value = []
        
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            result = service.generate_statement(1, "2025-10-15T23:59:59Z")
        
        assert result.cutoff_date == "2025-10-15T23:59:59Z"
        
        # Verify cutoff date was used in queries
        execute_calls = cursor.execute.call_args_list
        cutoff_date_found = False
        for call in execute_calls:
            if call[0] and "created_at_utc <= ?" in call[0][0]:
                cutoff_date_found = True
                assert "2025-10-15T23:59:59Z" in call[0][1]
        
        assert cutoff_date_found


class TestErrorHandlingIntegration:
    """Test error handling in realistic scenarios."""
    
    def test_database_connection_failure(self, service):
        """Test handling of database connection failure."""
        with patch('src.services.statement_service.get_db_connection') as mock_get_conn:
            mock_get_conn.side_effect = Exception("Database connection failed")
            
            with pytest.raises(Exception, match="Database connection failed"):
                service.generate_statement(1, "2025-10-31T23:59:59Z")
    
    def test_partial_data_scenarios(self, service):
        """Test scenarios with partial or missing data."""
        conn = Mock()
        cursor = Mock()
        conn.cursor.return_value = cursor
        
        # Mock scenario where some calculations return NULL
        cursor.fetchone.side_effect = [
            {"display_alias": "Test Associate", "home_currency": "EUR"},
            {"total_deposits": 1000.0, "signed_withdrawals": 0.0},
            {"current_holding_eur": 1000.0},
            {"should_hold_eur": None},    # No bet results yet
            {"profit_before_payout_eur": None},
        ]
        cursor.fetchall.return_value = []
        
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            result = service.generate_statement(1, "2025-10-31T23:59:59Z")

        # Should handle NULL gracefully
        assert result.should_hold_eur == Decimal("0.00")
        assert result.profit_before_payout_eur == Decimal("0.00")
        assert result.raw_profit_eur == Decimal("-1000.00")  # 0 - 1000
        assert result.delta_eur == Decimal("1000.00")      # 1000 - 0


class TestLargeDatasetPerformance:
    """Test performance with large datasets."""
    
    def test_large_transaction_history(self, service):
        """Test handling of large transaction history."""
        conn = Mock()
        cursor = Mock()
        conn.cursor.return_value = cursor
        
        # Mock calculation results
        cursor.fetchone.side_effect = [
            {"display_alias": "High Volume Associate", "home_currency": "EUR"},
            {"total_deposits": 50000.0, "signed_withdrawals": 0.0},
            {"current_holding_eur": 54500.0},
            {"should_hold_eur": 55000.0},
            {"profit_before_payout_eur": 0.0},
        ]
        
        # Mock large transaction list
        large_transaction_list = []
        for i in range(1000):
            large_transaction_list.append({
                "id": i,
                "type": "BET_RESULT",
                "amount_eur": "100.00",
                "created_at_utc": "2025-10-31T23:59:59Z"
            })
        
        def execute_side_effect(query, params=None):
            normalized = " ".join(query.split())
            if "SELECT id, type, amount_eur" in normalized:
                cursor.fetchall.return_value = large_transaction_list
            elif "FROM bookmakers" in normalized:
                cursor.fetchall.return_value = []
            else:
                cursor.fetchall.return_value = []

        cursor.execute.side_effect = execute_side_effect
        
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            transactions = service.get_associate_transactions(1, "2025-10-31T23:59:59Z")
        
        assert len(transactions) == 1000
        
        # Test that calculations still work with large values
        with patch('src.services.statement_service.get_db_connection', return_value=conn):
            result = service.generate_statement(1, "2025-10-31T23:59:59Z")
        assert result.net_deposits_eur == Decimal("50000.00")
        assert result.should_hold_eur == Decimal("55000.00")
        assert result.current_holding_eur == Decimal("54500.00")
        assert result.profit_before_payout_eur == Decimal("0.00")


class TestStatementExcelExport:
    """Test Excel export contains profit before payout metric."""

    def test_statement_excel_exports_allocations_and_totals(self, tmp_path):
        module = importlib.import_module("src.ui.pages.6_statements")
        original_dir = module.STATEMENT_EXPORT_DIR
        module.STATEMENT_EXPORT_DIR = tmp_path
        try:
            bookmakers = [
                BookmakerStatementRow(
                    bookmaker_name="CSV Bookie",
                    balance_eur=Decimal("250.00"),
                    deposits_eur=Decimal("200.00"),
                    withdrawals_eur=Decimal("0.00"),
                    balance_native=Decimal("250.00"),
                    native_currency="EUR",
                ),
                BookmakerStatementRow(
                    bookmaker_name="Second Bookie",
                    balance_eur=Decimal("150.00"),
                    deposits_eur=Decimal("175.00"),
                    withdrawals_eur=Decimal("25.00"),
                    balance_native=Decimal("150.00"),
                    native_currency="EUR",
                ),
            ]
            calc = StatementCalculations(
                associate_id=9,
                net_deposits_eur=Decimal("200.00"),
                should_hold_eur=Decimal("450.00"),
                current_holding_eur=Decimal("400.00"),
                fair_share_eur=Decimal("250.00"),
                profit_before_payout_eur=Decimal("250.00"),
                raw_profit_eur=Decimal("250.00"),
                delta_eur=Decimal("-50.00"),
                total_deposits_eur=Decimal("500.00"),
                total_withdrawals_eur=Decimal("300.00"),
                bookmakers=bookmakers,
                associate_name="CSV Tester",
                home_currency="EUR",
                cutoff_date="2025-10-31T23:59:59Z",
                generated_at="2025-11-01T00:00:00Z",
            )
            with patch.object(
                module.StatementService, "_calculate_multibook_delta", return_value=Decimal("0")
            ):
                excel_path = module.generate_statement_summary_excel(calc)
                workbook = load_workbook(excel_path)
                try:
                    worksheet = workbook.active
                    rows = [list(row or []) for row in worksheet.iter_rows(values_only=True)]
                finally:
                    workbook.close()

            assert rows[0] == ["Associate", "CSV Tester"]
            assert rows[1][0] == "As of (UTC)"
            assert rows[3][1].startswith("Totals in EUR.")

            header = [
                "Bookmaker",
                "Balance Native",
                "",
                "CCY",
                "Balance EUR",
                "CCY_EUR",
            ]
            header_index = rows.index(header)

            def to_decimal(value: str) -> Decimal:
                return Decimal(value.replace(",", "")) if value else Decimal("0")

            table_rows: List[List[str]] = []
            for row in rows[header_index + 1 :]:
                if not row:
                    break
                table_rows.append(row)

            assert len(table_rows) == 2
            assert table_rows[0][0] == "CSV Bookie"
            assert table_rows[0][2] == ""
            assert table_rows[0][3] == "EUR"
            assert table_rows[0][5] == "EURO"

            summary_start = rows.index([], header_index)
            summary_rows = [
                row for row in rows[summary_start + 1 :] if row and row[0] != "Footnote"
            ]
            summary_map = {row[0]: to_decimal(row[1]) for row in summary_rows}

            assert summary_map["Net Deposits (ND)"] == Decimal("200.00")
            assert summary_map["Fair Share (FS)"] == Decimal("250.00")
            assert summary_map["Imbalance (I'' = TB - YF)"] == Decimal("-50.00")
            assert summary_map["Exit Payout (-I'')"] == Decimal("50.00")
            assert summary_map["Multibook Delta"] == Decimal("0.00")
            assert summary_map["UTILE (YF - ND)"] == Decimal("250.00")
            assert any(row for row in rows if row and row[0] == "Footnote")
        finally:
            module.STATEMENT_EXPORT_DIR = original_dir

    def test_surebet_roi_export_handles_zero_stake(self, service, monkeypatch):
        calc = StatementCalculations(
            associate_id=3,
            net_deposits_eur=Decimal("100.00"),
            should_hold_eur=Decimal("150.00"),
            current_holding_eur=Decimal("160.00"),
            fair_share_eur=Decimal("40.00"),
            profit_before_payout_eur=Decimal("40.00"),
            raw_profit_eur=Decimal("50.00"),
            delta_eur=Decimal("10.00"),
            total_deposits_eur=Decimal("300.00"),
            total_withdrawals_eur=Decimal("200.00"),
            bookmakers=[],
            associate_name="ROI Tester",
            home_currency="EUR",
            cutoff_date="2025-11-30T23:59:59Z",
            generated_at="2025-12-01T00:00:00Z",
        )

        mock_cursor = Mock()
        mock_cursor.execute.return_value = mock_cursor
        mock_cursor.fetchall.return_value = [
            {
                "surebet_id": 101,
                "settled_at_utc": "2025-11-01T10:00:00Z",
                "associate_stake_eur": 300.0,
                "associate_profit_eur": 45.0,
                "group_stake_eur": 1200.0,
                "group_profit_eur": 120.0,
            },
            {
                "surebet_id": 102,
                "settled_at_utc": "2025-11-05T12:00:00Z",
                "associate_stake_eur": 0.0,
                "associate_profit_eur": 0.0,
                "group_stake_eur": 800.0,
                "group_profit_eur": 40.0,
            },
        ]

        mock_conn = Mock()
        mock_conn.cursor.return_value = mock_cursor
        mock_conn.close = Mock()
        monkeypatch.setattr(
            "src.services.statement_service.get_db_connection", lambda: mock_conn
        )

        export = service.export_surebet_roi_excel(
            associate_id=calc.associate_id,
            cutoff_date=calc.cutoff_date,
            calculations=calc,
        )

        workbook = load_workbook(io.BytesIO(export.content))
        try:
            worksheet = workbook.active
            rows = [list(row or []) for row in worksheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
        table_index = next(idx for idx, row in enumerate(rows) if row and row[0] == "Surebet ID")
        roi_rows = [row for row in rows[table_index + 1 :] if row and str(row[0]).isdigit()]

        assert len(roi_rows) == 2
        first_row = roi_rows[0]
        assert first_row[0] == "101"
        assert first_row[2] == "300.00"
        assert first_row[3] == "45.00"
        assert first_row[4] == "15.00%"
        assert first_row[6] == "120.00"
        assert first_row[7] == "10.00%"

        second_row = roi_rows[1]
        assert second_row[0] == "102"
        assert second_row[2] == "0.00"
        assert second_row[4] == ""
        assert second_row[7] == "5.00%"


class TestSettlementWorkflow:
    """Integration coverage for the Settle Associate Now workflow."""

    def test_settle_associate_now_zeroes_delta(self, temp_statement_db, monkeypatch):
        conn = sqlite3.connect(temp_statement_db)
        conn.row_factory = sqlite3.Row
        create_schema(conn)
        conn.execute(
            "INSERT INTO associates (id, display_alias, home_currency) VALUES (?, ?, ?)",
            (1, "Exit Tester", "EUR"),
        )
        conn.execute(
            """
            INSERT INTO ledger_entries (
                type, associate_id, bookmaker_id, amount_native, native_currency,
                fx_rate_snapshot, amount_eur, created_at_utc, created_by, note
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "DEPOSIT",
                1,
                None,
                "500.00",
                "EUR",
                "1.00",
                "500.00",
                "2025-10-01T00:00:00Z",
                "pytest",
                "initial funding",
            ),
        )
        conn.execute(
            """
            INSERT INTO ledger_entries (
                type, associate_id, bookmaker_id, amount_native, native_currency,
                fx_rate_snapshot, amount_eur, created_at_utc, created_by, note
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "BOOKMAKER_CORRECTION",
                1,
                None,
                "100.00",
                "EUR",
                "1.00",
                "100.00",
                "2025-10-15T00:00:00Z",
                "pytest",
                "manual adjustment",
            ),
        )
        conn.commit()
        conn.close()

        def _connect():
            connection = sqlite3.connect(temp_statement_db)
            connection.row_factory = sqlite3.Row
            return connection

        monkeypatch.setattr("src.services.statement_service.get_db_connection", _connect)
        monkeypatch.setattr(
            "src.services.funding_transaction_service.get_db_connection", _connect
        )

        service = StatementService()
        cutoff = "2025-10-31T23:59:59Z"

        calc_before = service.generate_statement(1, cutoff)
        assert calc_before.i_double_prime_eur == Decimal("100.00")

        result = service.settle_associate_now(
            1, cutoff, calculations=calc_before, created_by="pytest"
        )
        assert result.was_posted is True
        assert result.entry_type == "WITHDRAWAL"
        assert result.amount_eur == Decimal("100.00")
        assert result.delta_after == Decimal("0.00")

        calc_after = service.generate_statement(1, cutoff)
        assert calc_after.i_double_prime_eur == Decimal("0.00")



